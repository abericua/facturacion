import openpyxl
import os

# Ruta al archivo original
input_file_path = r"C:\Users\solpr\Desktop\Solpro\SOLPRO_2026_Sistema_Gestion_V6.xlsx"

def find_header_row(filepath):
    """Escanea las primeras filas para encontrar la fila de encabezado y la muestra."""
    if not os.path.exists(filepath):
        print(f"Error: No se encuentra el archivo en: {filepath}")
        return

    wb = openpyxl.load_workbook(filepath)
    sheet_name = "CALCULADORA"
    
    if sheet_name not in wb.sheetnames:
        print(f"Error: No se encontró la pestaña '{sheet_name}'.")
        return

    ws = wb[sheet_name]
    
    header_row_found = None
    header_row_num = -1

    # Escanear las primeras 10 filas
    for row_num in range(1, 11):
        row_values = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[row_num]]
        if "LÍNEA" in row_values or "COSTO USD" in row_values:
            header_row_found = ws[row_num]
            header_row_num = row_num
            break
    
    if header_row_found:
        headers = [(cell.column, cell.value) for cell in header_row_found]
        print(f"--- Fila de encabezado encontrada en la fila: {header_row_num} ---")
        for col_idx, header_name in headers:
            print(f"Columna {col_idx}: '{header_name}'")
    else:
        print("--- No se encontró una fila de encabezado con 'LÍNEA' o 'COSTO USD' en las primeras 10 filas. ---")


if __name__ == "__main__":
    find_header_row(input_file_path)