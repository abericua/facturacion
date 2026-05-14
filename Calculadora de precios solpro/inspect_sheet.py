
import openpyxl
import os

# Ruta al archivo original
input_file_path = r"C:\Users\solpr\Desktop\Solpro\SOLPRO_2026_Sistema_Gestion_V6.xlsx"

def inspect_sheet(filepath, sheet_name_to_inspect):
    """Inspecciona una hoja específica, listando encabezados y algunas filas de datos."""
    if not os.path.exists(filepath):
        print(f"Error: No se encuentra el archivo en: {filepath}")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    if sheet_name_to_inspect not in wb.sheetnames:
        print(f"Error: La pestaña '{sheet_name_to_inspect}' no existe en el archivo.")
        print(f"Pestañas disponibles: {wb.sheetnames}")
        return

    ws = wb[sheet_name_to_inspect]
    
    print(f"--- Inspeccionando las primeras 20 filas de la pestaña '{sheet_name_to_inspect}' ---")
    
    header_row_found = None
    header_row_num = -1

    # Buscar la fila de encabezado en las primeras 10 filas
    for row_num in range(1, 11):
        row_values = [cell.value for cell in ws[row_num]]
        if row_values and any(row_values): # Asegurarse de que no esté vacía
            header_row_num = row_num
            header_row_found = row_values
            print(f"\nPosible Fila de Encabezado encontrada en la Fila {header_row_num}:")
            print(header_row_found)
            break # Usar la primera fila no vacía como el encabezado probable
            
    if not header_row_found:
        print("No se encontró una fila de encabezado no vacía en las primeras 10 filas.")
        return

    # Imprimir las 10 filas de datos siguientes al encabezado
    print("\nMostrando las siguientes 10 filas de datos:")
    for row_num in range(header_row_num + 1, header_row_num + 11):
        row_values = [cell.value for cell in ws[row_num]]
        if any(row_values):
            print(f"Fila {row_num}: {row_values}")


if __name__ == "__main__":
    inspect_sheet(input_file_path, "BASE_DATOS")
