
import openpyxl
import os

# Ruta al archivo original
input_file_path = r"C:\Users\solpr\Desktop\Solpro\SOLPRO_2026_Sistema_Gestion_V6.xlsx"

def deep_inspect(filepath):
    """Inspecciona profundamente el archivo, listando hojas y volcando el contenido de 'CALCULADORA'."""
    if not os.path.exists(filepath):
        print(f"Error: No se encuentra el archivo en: {filepath}")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True) # data_only para ver valores de fórmulas antiguas
    
    print("--- Pestañas Disponibles en el Archivo ---")
    print(wb.sheetnames)
    print("-" * 40)

    sheet_name = "CALCULADORA"
    
    if sheet_name not in wb.sheetnames:
        print(f"Error: La pestaña '{sheet_name}' no existe en el archivo.")
        return

    ws = wb[sheet_name]
    
    print(f"--- Volcando las primeras 20 filas de la pestaña '{sheet_name}' ---")
    for row_num in range(1, 21):
        row_values = [cell.value for cell in ws[row_num]]
        # Filtrar filas completamente vacías para no saturar la salida
        if any(row_values):
            print(f"Fila {row_num}: {row_values}")

if __name__ == "__main__":
    deep_inspect(input_file_path)
