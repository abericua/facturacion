import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import json
from copy import copy

# --- CARGAR CONFIGURACIÓN ---
base_dir = os.path.dirname(__file__)
config_path = os.path.join(base_dir, 'config_solpro.json')
with open(config_path, 'r') as f:
    config = json.load(f)
    params = config['parametros']

# Rutas Relativas
input_file_path = os.path.join(base_dir, config['archivos']['excel_gestion'])
output_file_name = config['archivos']['excel_gestion'].replace(".xlsx", "_V7_CONFIG.xlsx")
output_file_path = os.path.join(base_dir, output_file_name)

def setup_parametros_sheet(wb):
    """Crea y configura la pestaña PARAMETROS."""
    sheet_name = "PARAMETROS"
    if sheet_name in wb.sheetnames: del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)
    
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    
    params_data = {
        "A1": "Variable", "B1": "Valor", "A2": "Dólar Mercado", "B2": 7350,
        "A3": "Buffer Piso (Protección Diaria)", "B3": params['buffer_piso'], 
        "A4": "Buffer Techo (Protección Crédito)", "B4": params['buffer_techo'],
        "A5": "Margen Deseado", "B5": params['margen_deseado_pct'] / 100, 
        "A6": "Comisión QR/Tarjeta", "B6": params['comision_qr_pct'] / 100,
        "A7": "Factor Redondeo Gs", "B7": params['factor_redondeo_gs'],
        "A8": "Umbral Redondeo", "B8": params['umbral_redondeo_gs'],
        "A9": "Cálculos Automáticos",
        "A10": "BANDA PISO", "B10": "=B2+B3", 
        "A11": "BANDA TECHO", "B11": "=B2+B4"
    }

    for cell, value in params_data.items(): ws[cell] = value
    for cell in ["A1", "B1", "A9"]:
        ws[cell].font, ws[cell].fill = header_style, header_fill
    ws["B5"].style, ws["B6"].style = percent_style, percent_style
    for cell in ["B2", "B3", "B4", "B10", "B11", "B7", "B8"]: ws[cell].number_format = '#,##0'
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    print("Pestaña 'PARAMETROS' configurada.")

def process_calculadora_sheet_debug(wb):
    """Añade columnas de depuración con fórmulas que siguen la política de Solpro."""
    sheet_name = "CALCULADORA"
    if sheet_name not in wb.sheetnames: return

    ws = wb[sheet_name]
    HEADER_ROW, DATA_START_ROW = 4, 6
    STYLE_REF_CELL = 'C4'
    
    # Encontrar columnas clave
    col_linea_idx = None
    for cell in ws[HEADER_ROW]:
        if "TIPO" in str(cell.value): col_linea_idx = cell.column; break
            
    if not col_linea_idx:
        print("Error: No se encontró la columna de TIPO."); return

    next_col = ws.max_column + 1
    col_costo_final = get_column_letter(next_col)
    col_p_contado = get_column_letter(next_col + 1)
    col_p_qr = get_column_letter(next_col + 2)

    new_headers = {
        col_costo_final: "COSTO FINAL (debug)",
        col_p_contado: "P. CONTADO (SOLPRO)",
        col_p_qr: "P. QR (SOLPRO)"
    }
    
    ref_cell = ws[STYLE_REF_CELL]
    for col_letter, title in new_headers.items():
        cell = ws[f'{col_letter}{HEADER_ROW}']
        cell.value, cell.font, cell.fill, cell.alignment = title, copy(ref_cell.font), copy(ref_cell.fill), copy(ref_cell.alignment)

    ref_banda_piso = "PARAMETROS!$B$10"
    ref_margen = "PARAMETROS!$B$5"
    ref_comision_qr = "PARAMETROS!$B$6"
    ref_redondeo = "PARAMETROS!$B$7"

    for row in range(DATA_START_ROW, ws.max_row + 1):
        if ws[f'A{row}'].value is None: continue

        # Fórmula de Precio Contado Solpro: ROUNDUP((Costo * Banda) / (1-Margen), -4) - 1000
        # (Ajustado para que termine en 9000)
        costo_usd_cell = f"D{row}" # Asumiendo columna D es Costo USD
        ws[f'{col_costo_final}{row}'] = f"={costo_usd_cell}*{ref_banda_piso}"
        
        # Lógica de Redondeo Menor Solpro en Excel
        formula_contado = f"=(ROUNDUP(({col_costo_final}{row})/(1-{ref_margen}), -4) - 1000)"
        ws[f'{col_p_contado}{row}'] = formula_contado
        
        # Lógica de QR (+4%)
        ws[f'{col_p_qr}{row}'] = f"=(ROUNDUP({col_p_contado}{row}/(1-{ref_comision_qr}), -4) - 1000)"

    print("Columnas Solpro añadidas con éxito.")

def main():
    if not os.path.exists(input_file_path):
        print(f"Error: El archivo '{input_file_path}' no existe."); return
    wb = openpyxl.load_workbook(input_file_path)
    setup_parametros_sheet(wb)
    process_calculadora_sheet_debug(wb)
    try:
        wb.save(output_file_path)
        print(f"\n¡Éxito! Archivo guardado en: {output_file_path}")
    except PermissionError:
        print("\nError: Cierra el archivo Excel antes de ejecutar el script.")

if __name__ == "__main__":
    main()
